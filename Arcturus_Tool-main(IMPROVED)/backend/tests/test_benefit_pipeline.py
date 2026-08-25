"""
Offline verification harness for the Business Benefit pipeline.

Runs with NO network access and NO API keys. It uses HTML fixtures shaped like
Oracle readiness feature pages and a scripted stub model, so it verifies the
pipeline's control flow and quality gate, not live output quality.

Run from the repository root:

    python backend/tests/test_benefit_pipeline.py

Optional: pass a path to an OQUAT report to also run the quality gate over real
generated output:

    python backend/tests/test_benefit_pipeline.py path/to/OQUAT_Report.xlsx
"""

import asyncio
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
BACKEND = os.path.dirname(HERE)
sys.path.insert(0, BACKEND)

from bs4 import BeautifulSoup  # noqa: E402

from services import ai_enricher as ae  # noqa: E402

FIXTURES = os.path.join(HERE, "fixtures")


def _soup(name):
    with open(os.path.join(FIXTURES, name), encoding="utf-8") as fh:
        return BeautifulSoup(fh.read(), "html.parser")


def test_extraction_with_h1():
    """Pages that DO have an <h1> keep working, and the benefit section is captured."""
    soup = _soup("feature_page_with_h1.html")
    body = ae.extract_feature_body_text(soup)
    benefit = ae.extract_oracle_benefit_section(soup)

    assert "Redwood experience to configure collaboration messaging" in body
    # The Steps-to-Enable config table must NOT leak into the description.
    assert "This value is in hours" not in body
    assert "onboard a new trading partner" in benefit

    print("[PASS] with-h1 page: body extracted, config table excluded, "
          "Oracle benefit section captured")


def test_extraction_without_h1():
    """The regression case: no <h1>, description nested in a wrapper div."""
    soup = _soup("feature_page_no_h1.html")

    old_intro = ae.extract_intro_text(soup)
    old_section = ae.extract_section_text(
        soup, ["overview", "description", "feature summary", "details"]
    )
    new_body = ae.extract_feature_body_text(soup)

    assert old_intro == "", "expected the old sibling-walk path to return nothing"
    assert old_section == "", "expected the old heading-name path to return nothing"
    assert "define rules that modify inbound B2B messages" in new_body
    assert "Leave the ID cell blank" not in new_body

    print("[PASS] no-h1 page: old paths returned '', new path recovered the "
          "description without pulling in Steps-to-Enable text")


def test_quality_gate():
    """The gate must reject exactly the failure modes seen in shipped output."""
    source = ("Using License Plate Number Real Time Subject Area in OTBI, you can now "
              "create personalized reports on your LPNs and the contents within LPNs.")

    bad = [
        "Provides this value is in hours.",
        "Improves leave the ID cell blank.",
        "Provides the ability to create personalized reports for your LPNs.",
        "Improves operational efficiency for warehouse management.",
        ("Improves using License Plate Number Real Time Subject Area in OTBI, you can "
         "now create personalized reports on your LPNs and the contents within LPNs."),
    ]
    accepted, reasons = ae.validate_benefit_lines(bad, source)
    assert accepted == [], f"expected all rejected, got {accepted}"
    assert len(reasons) == len(bad)

    good = [
        ("Cuts the reporting turnaround for warehouse supervisors investigating stock "
         "discrepancies, so container-level questions are answered from self-service "
         "analysis rather than a raised IT request."),
        ("Reduces dependence on central reporting teams by letting inventory analysts "
         "assemble their own container-content views, shortening the cycle between a "
         "question being asked and a decision being made."),
    ]
    accepted, reasons = ae.validate_benefit_lines(good, source)
    assert len(accepted) == 2, f"expected both accepted, reasons={reasons}"

    print("[PASS] quality gate: 5/5 known-bad rejected, 2/2 known-good accepted")


def test_retry_recovers():
    """Scenario A: model returns junk twice, then clean output. Must not fall back."""
    soup = _soup("feature_page_no_h1.html")

    async def fake_fetch(_url):
        return {
            "full_text": ae.clean_text(soup.get_text(" "))[:50000],
            "intro_source": ae.extract_intro_text(soup),
            "description_source": ae.extract_feature_body_text(soup),
            "impact_source": "",
            "steps_source": "",
            "access_source": "",
            "benefit_source": ae.extract_oracle_benefit_section(soup),
        }

    script = [
        "Provides the ability to define processing rules for inbound B2B messages.\n"
        "Supports message modification without development.",

        "1. This feature enables organizations to modify messages.\n"
        "2. Rules are maintained in a spreadsheet that business users can update directly.",

        "Cuts dependence on custom development by letting business users adjust inbound "
        "B2B message rules directly to meet enterprise collaboration requirements.\n"
        "Keeps inbound message rules maintainable by business users, making enterprise-specific "
        "message changes easier to apply without custom development.",
    ]
    calls = {"n": 0}

    async def fake_ai(_prompt):
        i = min(calls["n"], len(script) - 1)
        calls["n"] += 1
        return script[i]

    orig_fetch, orig_ai, orig_delay = (
        ae.fetch_oracle_detail_text, ae.call_ai, ae.REQUEST_DELAY_SECONDS
    )
    ae.fetch_oracle_detail_text, ae.call_ai, ae.REQUEST_DELAY_SECONDS = (
        fake_fetch, fake_ai, 0
    )
    try:
        feature = {
            "title": "Redwood: Define Processing Rules for B2B Messages",
            "url": "https://docs.oracle.com/example.htm",
        }
        out = asyncio.run(ae.enrich_feature(dict(feature), 1, 1))
    finally:
        ae.fetch_oracle_detail_text, ae.call_ai, ae.REQUEST_DELAY_SECONDS = (
            orig_fetch, orig_ai, orig_delay
        )

    assert out["benefit_origin"] == "AI_RETRY_3", out["benefit_origin"]
    assert "Cuts dependence on custom development" in out["business_benefit"]
    print("[PASS] retry loop: recovered on attempt 3, benefit_origin=AI_RETRY_3")


def test_total_failure_never_splices_documentation():
    """Scenario B: every attempt fails. Fallback must copy nothing from the page."""
    soup = _soup("feature_page_no_h1.html")
    page_text = ae.clean_text(soup.get_text(" "))

    async def fake_fetch(_url):
        return {
            "full_text": page_text[:50000],
            "intro_source": "",
            "description_source": ae.extract_feature_body_text(soup),
            "impact_source": "",
            "steps_source": "",
            "access_source": "",
            "benefit_source": "",
        }

    async def always_fail(_prompt):
        raise RuntimeError("All AI providers failed -> groq: 429 | gemini: no key")

    orig_fetch, orig_ai, orig_delay = (
        ae.fetch_oracle_detail_text, ae.call_ai, ae.REQUEST_DELAY_SECONDS
    )
    ae.fetch_oracle_detail_text, ae.call_ai, ae.REQUEST_DELAY_SECONDS = (
        fake_fetch, always_fail, 0
    )
    try:
        feature = {
            "title": "Redwood: Define Processing Rules for B2B Messages",
            "url": "https://docs.oracle.com/example.htm",
        }
        out = asyncio.run(ae.enrich_feature(dict(feature), 1, 1))
    finally:
        ae.fetch_oracle_detail_text, ae.call_ai, ae.REQUEST_DELAY_SECONDS = (
            orig_fetch, orig_ai, orig_delay
        )

    benefit = out["business_benefit"]
    assert out["benefit_origin"] == "FALLBACK"
    assert benefit.strip(), "fallback must never emit an empty Business Benefit"

    # The core regression: no sentence from the page may appear in the output.
    for sentence in page_text.split(". "):
        sentence = sentence.strip().rstrip(".")
        if len(sentence.split()) >= 5:
            assert sentence.lower() not in benefit.lower(), (
                f"fallback copied source text: {sentence!r}"
            )

    # CHANGED: the reviewer note used to be written INTO the benefit cell.
    # That text is customer-facing (it lands in the client deck), so the
    # signal now goes to the run log instead. What the cell must contain is a
    # feature-specific statement, never an internal process note.
    assert "not generated by the AI pipeline" not in benefit
    assert "requires functional review" not in benefit

    # Feature-specific: the fallback must reference this feature's own domain.
    assert "B2B" in benefit or "message" in benefit.lower(), (
        f"fallback is not feature-specific: {benefit!r}"
    )
    print("[PASS] total-failure fallback: no source sentence copied, cell "
          "non-empty, feature-specific, no internal note leaked to the deck")


def test_old_fallback_reproduces_reported_defect():
    """
    Documents the original defect mechanism. Reconstructs the old sentence-splicing
    behaviour and shows it emits Steps-to-Enable fragments when description is empty.
    """
    soup = _soup("feature_page_no_h1.html")
    page_text = ae.clean_text(soup.get_text(" "))

    # Old code path: description_source == "" -> fell through to full_text.
    assert ae.extract_intro_text(soup) == ""
    assert "Leave the ID cell blank" in page_text

    # New fallback given the same inputs must not surface that fragment.
    new_out = ae.generate_benefit_fallback(
        "Redwood: Define Processing Rules for B2B Messages", "", page_text
    )
    assert "leave the id cell blank" not in new_out.lower()
    print("[PASS] regression: setup-table fragments can no longer reach the output")


def report_gate_over_real_output(path):
    """Optional: run the gate across an existing OQUAT report."""
    import openpyxl

    wb = openpyxl.load_workbook(path)
    ws = wb["Release Notes"]
    header = [c.value for c in ws[1]]
    bi = header.index("Business Benefit")
    di = header.index("Description")

    rows = passed = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows += 1
        lines = [l for l in (row[bi] or "").split("\n") if l.strip()]
        accepted, _ = ae.validate_benefit_lines(lines, row[di] or "")
        if len(accepted) >= 2:
            passed += 1

    print(f"\n[GATE] {os.path.basename(path)}: {passed}/{rows} rows pass the quality "
          f"gate; {rows - passed} would be regenerated")


def test_exactly_two_benefits_are_required():
    source = "License Plate Number Real Time Subject Area in OTBI provides reporting and analysis of LPNs across inventory, receiving and shipping."
    title = "Create LPN Reports Using the License Plate Number Real Time Subject Area in OTBI"
    one = ["Self-service LPN reporting gives inventory teams direct access to container-level transaction analysis across receiving, shipping, and inventory processes."]
    accepted, _ = ae.validate_benefit_lines(one, source, title=title, evidence_blob=source.lower())
    assert len(accepted) == 1

    two = [
        "Self-service LPN reporting gives inventory teams direct access to container-level transaction analysis across receiving, shipping, and inventory processes.",
        "Personalized LPN views make it easier to examine container contents and transaction information within the supported inventory, receiving, and shipping workflows.",
    ]
    accepted, reasons = ae.validate_benefit_lines(two, source, title=title, evidence_blob=source.lower())
    assert len(accepted) == 2, reasons
    print("[PASS] exact-two policy: the pipeline distinguishes one statement from the required two")


def test_prompt_contains_complete_oracle_dossier_and_exact_two_rule():
    captured = {}

    async def fake_call_ai(prompt):
        captured["prompt"] = prompt
        return "Benefit one is grounded in LPN reporting and analysis.\nBenefit two is grounded in LPN transaction visibility."

    original = ae.call_ai
    ae.call_ai = fake_call_ai
    try:
        source = {
            "full_text": "UNIQUE_COMPLETE_ORACLE_FEATURE_DOSSIER_12345",
            "benefit_source": "Oracle says users can create personalized LPN reports.",
            "impact_source": "LPN usage is required for business impact.",
            "steps_source": "Configure the required setup.",
        }
        import asyncio
        asyncio.run(ae.generate_benefit_ai(
            "Create LPN Reports Using the License Plate Number Real Time Subject Area in OTBI",
            "Users can create personalized LPN reports.",
            "LPN reporting overview.",
            source,
        ))
    finally:
        ae.call_ai = original

    prompt = captured["prompt"]
    assert "Complete Oracle Feature Information" in prompt
    assert "UNIQUE_COMPLETE_ORACLE_FEATURE_DOSSIER_12345" in prompt
    assert "EXACTLY TWO" in prompt
    assert "do not return\nthree or more" in prompt
    print("[PASS] HR grounding prompt: complete Oracle dossier is supplied and exactly two benefits are required")


if __name__ == "__main__":
    test_extraction_with_h1()
    test_extraction_without_h1()
    test_quality_gate()
    test_retry_recovers()
    test_total_failure_never_splices_documentation()
    test_old_fallback_reproduces_reported_defect()
    test_exactly_two_benefits_are_required()
    test_prompt_contains_complete_oracle_dossier_and_exact_two_rule()

    for arg in sys.argv[1:]:
        report_gate_over_real_output(arg)

    print("\nAll offline checks passed.")
