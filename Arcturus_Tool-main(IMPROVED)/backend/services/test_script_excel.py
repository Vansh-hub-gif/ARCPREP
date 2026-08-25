import os
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="008063", end_color="008063", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
WRAP_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

COLUMN_WIDTHS = {
    "A": 22,  # Expanded slightly for variable script numbers
    "B": 25,
    "C": 22,
    "D": 25,
    "E": 45,
    "F": 65,
}

def generate_test_script_excel(
    test_scripts: List[Dict[str, str]],
    output_path: str = "outputs/oquat_test_scripts.xlsx",
) -> str:
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "UAT Test Scripts"

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    headers = [
        "Test Case ID",
        "L1 Process Family",
        "L2 Process Area",
        "L3 Process",
        "Feature Name",
        "Description",
    ]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_num, script in enumerate(test_scripts, start=2):
        for col_num, header in enumerate(headers, start=1):
            value = script.get(header, "")
            cell = ws.cell(row=row_num, column=col_num, value=value)
            
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGNMENT
            cell.border = THIN_BORDER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(test_scripts) + 1}"

    wb.save(output_path)
    return output_path